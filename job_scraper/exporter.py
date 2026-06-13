import csv
import json
from typing import List, Dict
from datetime import datetime
from pathlib import Path


class JobExporter:
    """Exports job data to various formats."""
    
    def __init__(self, output_dir: str = "."):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_csv(self, jobs: List[Dict], filename: str = None) -> str:
        """Export jobs to CSV file."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"jobs_{timestamp}.csv"
        
        filepath = self.output_dir / filename
        
        # Flatten nested data
        flat_jobs = []
        for job in jobs:
            flat_job = {
                "title": job.get("title", ""),
                "company": job.get("company", ""),
                "location": job.get("location", ""),
                "source": job.get("source", ""),
                "url": job.get("url", ""),
                "skills": " | ".join(job.get("skills", [])),
                "experience": job.get("experience", ""),
                "requirements": " | ".join(job.get("requirements", [])),
                "responsibilities": " | ".join(job.get("responsibilities", [])),
                "salary": job.get("salary", "Not specified"),
                "remote": job.get("remote", "Unknown"),
                "scraped_at": datetime.now().isoformat(),
            }
            flat_jobs.append(flat_job)
        
        if flat_jobs:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=flat_jobs[0].keys())
                writer.writeheader()
                writer.writerows(flat_jobs)
        
        return str(filepath)
    
    def export_excel(self, jobs: List[Dict], filename: str = None) -> str:
        """Export jobs to Excel file (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("Install openpyxl: pip install openpyxl")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"jobs_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Listings"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Headers
        headers = ["#", "Title", "Company", "Location", "Source", "Skills", 
                   "Experience", "Requirements", "URL", "Scraped At"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, job in enumerate(jobs, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=job.get("title", ""))
            ws.cell(row=row_idx, column=3, value=job.get("company", ""))
            ws.cell(row=row_idx, column=4, value=job.get("location", ""))
            ws.cell(row=row_idx, column=5, value=job.get("source", ""))
            ws.cell(row=row_idx, column=6, value=" | ".join(job.get("skills", [])))
            ws.cell(row=row_idx, column=7, value=job.get("experience", ""))
            ws.cell(row=row_idx, column=8, value=" | ".join(job.get("requirements", [])))
            ws.cell(row=row_idx, column=9, value=job.get("url", ""))
            ws.cell(row=row_idx, column=10, value=datetime.now().isoformat())
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(filepath)
        return str(filepath)
    
    def export_json(self, jobs: List[Dict], filename: str = None) -> str:
        """Export jobs to JSON file."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"jobs_{timestamp}.json"
        
        filepath = self.output_dir / filename
        
        export_data = {
            "exported_at": datetime.now().isoformat(),
            "total_jobs": len(jobs),
            "jobs": jobs
        }
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2, default=str)
        
        return str(filepath)
    
    def export_markdown(self, jobs: List[Dict], filename: str = None) -> str:
        """Export jobs to Markdown file."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"jobs_{timestamp}.md"
        
        filepath = self.output_dir / filename
        
        content = f"# Agentic AI Engineer Job Listings\n\n"
        content += f"*Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n\n"
        content += f"**Total Jobs Found: {len(jobs)}**\n\n---\n\n"
        
        for i, job in enumerate(jobs, 1):
            content += f"## {i}. {job.get('title', 'Unknown')}\n\n"
            content += f"- **Company:** {job.get('company', 'Unknown')}\n"
            content += f"- **Location:** {job.get('location', 'Not specified')}\n"
            content += f"- **Source:** {job.get('source', 'Unknown')}\n"
            content += f"- **Experience:** {job.get('experience', 'Not specified')}\n"
            
            if job.get("skills"):
                content += f"- **Skills:** {', '.join(job['skills'])}\n"
            
            if job.get("requirements"):
                content += f"\n### Requirements\n"
                for req in job["requirements"][:5]:
                    content += f"- {req}\n"
            
            if job.get("url"):
                content += f"\n[View Job Posting]({job['url']})\n"
            
            content += "\n---\n\n"
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        
        return str(filepath)
